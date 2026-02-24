from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable
import math
import re
import unicodedata
import json

import pandas as pd

EXCEL_NAME = "DASHBOARD MAIORES ENTREGAS.xlsx"
OUTPUT_NAME = "index.html"
PHOTO_DIR = Path("assets") / "colaboradores"
PHOTO_INPUT_DIR = Path("assets") / "fotos_colaboradores"
DEFAULT_SORT_COLUMNS = ["valor", "entregas", "peso"]
COLAB_SORT_COLUMNS = ["peso", "entregas", "valor"]
DEFAULT_RANKING_TEXT = "Ranking baseado em valor total, entregas e peso."
COLAB_RANKING_TEXT = "Ranking baseado em peso total e entregas. Valor exibido apenas para consulta."


def format_number(value: float, decimals: int = 0) -> str:
    if pd.isna(value):
        return "-"
    formatted = f"{value:,.{decimals}f}"
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    if decimals == 0:
        return formatted.split(",")[0]
    return formatted


def format_percentage(value: float) -> str:
    if pd.isna(value):
        return "-"
    return f"{value * 100:.0f}%".replace(".", ",")


def format_quantity(value: float) -> str:
    if pd.isna(value):
        return "-"
    if float(value).is_integer():
        return format_number(value, 0)
    return format_number(value, 1)


def format_compact_number(value: float, decimals: int = 1) -> str:
    if pd.isna(value):
        return "-"
    number = float(value)
    abs_number = abs(number)
    if abs_number >= 1_000_000:
        compact = format_number(number / 1_000_000, decimals).rstrip("0").rstrip(",")
        return f"{compact} mi"
    if abs_number >= 1_000:
        compact = format_number(number / 1_000, decimals).rstrip("0").rstrip(",")
        return f"{compact} mil"
    if number.is_integer():
        return format_number(number, 0)
    return format_number(number, min(decimals, 2))


def responsive_text(full_text: str, compact_text: str) -> str:
    if full_text == compact_text:
        return full_text
    return f'<span class="value-full">{full_text}</span><span class="value-compact">{compact_text}</span>'


def get_initials(nome: str) -> str:
    partes = [p for p in nome.split() if p]
    if not partes:
        return "--"
    if len(partes) == 1:
        return partes[0][0].upper()
    return (partes[0][0] + partes[-1][0]).upper()


def slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^A-Za-z0-9]+", "_", normalized)
    normalized = normalized.strip("_").lower()
    return normalized or "colaborador"


def normalize_key(value: str) -> str:
    return slugify(value)


def lookup_photo(name: str | None, photo_map: dict[str, str]) -> str | None:
    if not name:
        return None
    def normalize_spaces(text: str) -> str:
        return " ".join(text.split())
    candidates = [
        name.strip().lower(),
        normalize_spaces(name).lower(),
        normalize_key(name),
    ]
    for key in candidates:
        if key in photo_map:
            return photo_map[key]
    return None


def collect_photos(workbook_path: Path) -> dict[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    mapping: dict[str, str] = {}
    try:
        wb = load_workbook(workbook_path, data_only=True)
    except Exception:
        return {}

    photo_dir = workbook_path.parent / PHOTO_DIR
    photo_dir.mkdir(parents=True, exist_ok=True)

    for sheet_name in wb.sheetnames[1:]:
        ws = wb[sheet_name]
        images = getattr(ws, "_images", [])
        if not images:
            continue

        img = images[0]
        data = img._data()
        ext = getattr(img, "format", "png") or "png"
        ext = ext.lower()
        if ext == "jpeg":
            ext = "jpg"
        filename = f"{slugify(sheet_name)}.{ext}"
        file_path = photo_dir / filename
        try:
            file_path.write_bytes(data)
        except OSError:
            continue

        rel_path = file_path.relative_to(workbook_path.parent).as_posix()
        key_variants = {
            sheet_name.strip().lower(),
            normalize_key(sheet_name),
        }
        for key in key_variants:
            mapping[key] = rel_path

    return mapping


def load_local_photos(base_path: Path) -> dict[str, str]:
    """Carrega fotos da pasta local PHOTO_INPUT_DIR e mapeia por nome/slug."""
    mapping: dict[str, str] = {}
    base_path.mkdir(parents=True, exist_ok=True)
    for file in base_path.glob("*"):
        if not file.is_file():
            continue
        ext = file.suffix.lower()
        if ext not in {".jpg", ".jpeg", ".png", ".webp"}:
            continue
        stem = file.stem
        rel = file.relative_to(Path(__file__).resolve().parent).as_posix()
        base_name = stem.strip()
        keys = {
            base_name.lower(),
            " ".join(base_name.split()).lower(),
            normalize_key(base_name),
        }
        for key in keys:
            mapping[key] = rel
    return mapping


def load_planilha(path: Path) -> pd.DataFrame:
    def parse_number(value: object) -> float:
        if pd.isna(value):
            return math.nan
        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip()
        if not text:
            return math.nan

        # Mantem apenas caracteres numericos e separadores.
        text = re.sub(r"[^0-9,.\-]", "", text)
        if not text:
            return math.nan

        # Formato pt-BR: 1.234,56 -> 1234.56
        if "," in text:
            text = text.replace(".", "").replace(",", ".")
        elif text.count(".") > 1:
            # Remove separadores de milhar quando houver mais de um ponto.
            parts = text.split(".")
            text = "".join(parts[:-1]) + "." + parts[-1]

        try:
            return float(text)
        except ValueError:
            return math.nan

    def read_excel(skip_rows: int) -> pd.DataFrame:
        try:
            return pd.read_excel(path, skiprows=skip_rows)
        except Exception:
            return pd.DataFrame()

    df = read_excel(2)
    if df.empty or "DATA" not in df.columns:
        df = read_excel(0)
    if df.empty:
        raise ValueError("Nao foi possivel ler a planilha de entregas.")

    df = df.rename(
        columns={
            "DATA": "data",
            "MOTORISTA": "motorista",
            "AJUDANTE": "ajudante_1",
            "AJUDANTE.1": "ajudante_2",
            "CIDADE": "cidade",
            "ENTREGAS": "entregas",
            "PESO": "peso",
            "VALOR": "valor",
            "CLIENTE": "cliente",
            "PLACA": "placa",
        }
    )

    df["motorista"] = df["motorista"].apply(lambda nome: nome.strip() if isinstance(nome, str) else nome)
    df["motorista"] = df["motorista"].fillna("Sem motorista")

    if "cidade" not in df.columns:
        df["cidade"] = None
    df["cidade"] = df["cidade"].apply(lambda cidade: cidade.strip() if isinstance(cidade, str) else None).fillna(
        "Cidade nao informada"
    )

    if "cliente" not in df.columns:
        df["cliente"] = None
    df["cliente"] = df["cliente"].apply(lambda cliente: cliente.strip() if isinstance(cliente, str) else None).fillna(
        "Cliente nao informado"
    )

    if "placa" not in df.columns:
        df["placa"] = None
    df["placa"] = (
        df["placa"]
        .apply(lambda placa: placa.strip() if isinstance(placa, str) else None)
        .apply(lambda placa: re.sub(r"\s*-\s*$", "", placa) if isinstance(placa, str) else placa)
        .replace("", None)
        .fillna("Placa nao informada")
    )

    if "valor" not in df.columns:
        df["valor"] = 0
    if "entregas" not in df.columns:
        df["entregas"] = 1

    df["data"] = pd.to_datetime(df["data"], errors="coerce", dayfirst=True)
    for col in ("entregas", "peso", "valor"):
        df[col] = df[col].apply(parse_number).fillna(0)
    df["entregas"] = df["entregas"].where(df["entregas"] > 0, 1)

    df = df.dropna(how="all", subset=["data", "motorista", "entregas", "peso"])
    return df


def resumir_colaboradores(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    motoristas = (
        df.dropna(subset=["motorista"])
        .groupby("motorista", dropna=True)[["entregas", "peso", "valor"]]
        .sum()
        .sort_values(COLAB_SORT_COLUMNS, ascending=False)
        .reset_index()
        .rename(columns={"motorista": "colaborador"})
    )

    ajudantes_registros: list[dict[str, float]] = []
    for _, row in df.iterrows():
        entregas = row["entregas"]
        peso = row["peso"]
        valor = row["valor"]
        nomes: list[str] = []
        for col in ("ajudante_1", "ajudante_2"):
            nome = row.get(col)
            if isinstance(nome, str):
                nome_limpo = nome.strip()
                if nome_limpo and nome_limpo.lower() != "nan":
                    nome_normalizado = " ".join(parte.capitalize() for parte in nome_limpo.split())
                    nomes.append(nome_normalizado)

        if not nomes:
            continue

        fator = 1 / len(nomes)
        entregas_cota = entregas * fator
        peso_cota = peso * fator
        valor_cota = valor * fator

        for nome in nomes:
            ajudantes_registros.append(
                {"colaborador": nome, "entregas": entregas_cota, "peso": peso_cota, "valor": valor_cota}
            )

    if ajudantes_registros:
        ajudantes_df = pd.DataFrame(ajudantes_registros)
        ajudantes = (
            ajudantes_df.groupby("colaborador")[["entregas", "peso", "valor"]]
            .sum()
            .sort_values(COLAB_SORT_COLUMNS, ascending=False)
            .reset_index()
        )
    else:
        ajudantes = pd.DataFrame(columns=["colaborador", "entregas", "peso", "valor"])

    return motoristas, ajudantes


def dividir_ajudantes_por_linha(row: pd.Series, *, chave_extra: dict[str, object] | None = None) -> list[dict[str, object]]:
    ajudantes_registros: list[dict[str, object]] = []
    entregas = row.get("entregas", 0)
    peso = row.get("peso", 0)
    valor = row.get("valor", 0)
    extras = chave_extra or {}
    nomes: list[str] = []
    for col in ("ajudante_1", "ajudante_2"):
        nome = row.get(col)
        if isinstance(nome, str):
            nome_limpo = nome.strip()
            if nome_limpo and nome_limpo.lower() != "nan":
                nome_normalizado = " ".join(parte.capitalize() for parte in nome_limpo.split())
                nomes.append(nome_normalizado)

    if not nomes:
        return ajudantes_registros

    fator = 1 / len(nomes)
    entregas_cota = entregas * fator
    peso_cota = peso * fator
    valor_cota = valor * fator

    for nome in nomes:
        registro = {
            "colaborador": nome,
            "entregas": entregas_cota,
            "peso": peso_cota,
            "valor": valor_cota,
        }
        registro.update(extras)
        ajudantes_registros.append(registro)
    return ajudantes_registros


def ranking_motorista_por(df: pd.DataFrame, chave: str) -> pd.DataFrame:
    agrupado = (
        df.groupby([chave, "motorista"])[["entregas", "peso", "valor"]]
        .sum()
        .sort_values(COLAB_SORT_COLUMNS, ascending=False)
        .reset_index()
        .rename(columns={"motorista": "colaborador"})
    )
    agrupado["colaborador"] = agrupado["colaborador"].fillna("Sem motorista").apply(lambda x: str(x).strip().title())
    agrupado[chave] = agrupado[chave].fillna(f"{chave} nao informado").apply(lambda x: str(x).strip().title())
    agrupado["colaborador"] = agrupado["colaborador"] + " — " + agrupado[chave]
    return agrupado[["colaborador", "entregas", "peso", "valor"]]


def ranking_ajudante_por(df: pd.DataFrame, chave: str) -> pd.DataFrame:
    registros: list[dict[str, object]] = []
    for _, row in df.iterrows():
        chave_val = row.get(chave)
        chave_val = chave_val.strip().title() if isinstance(chave_val, str) else f"{chave} nao informado"
        registros.extend(dividir_ajudantes_por_linha(row, chave_extra={chave: chave_val}))

    if not registros:
        return pd.DataFrame(columns=["colaborador", "entregas", "peso", "valor"])

    agrupado = (
        pd.DataFrame(registros)
        .groupby(["colaborador", chave])[["entregas", "peso", "valor"]]
        .sum()
        .sort_values(COLAB_SORT_COLUMNS, ascending=False)
        .reset_index()
    )
    agrupado["colaborador"] = agrupado["colaborador"].apply(lambda x: str(x).strip().title())
    agrupado["colaborador"] = agrupado["colaborador"] + " — " + agrupado[chave]
    return agrupado[["colaborador", "entregas", "peso", "valor"]]


def resumir_clientes(df: pd.DataFrame) -> pd.DataFrame:
    base = df.copy()
    base["cliente"] = base["cliente"].fillna("Cliente nao informado")
    base["cliente"] = base["cliente"].apply(
        lambda nome: nome.strip() if isinstance(nome, str) and nome.strip() else "Cliente nao informado"
    )
    return (
        base.groupby("cliente")[["entregas", "peso", "valor"]]
        .sum()
        .sort_values(["valor", "entregas", "peso"], ascending=False)
        .reset_index()
        .rename(columns={"cliente": "colaborador"})
    )


def resumir_cidades(df: pd.DataFrame) -> pd.DataFrame:
    base = df.copy()
    base["cidade"] = base["cidade"].fillna("Cidade nao informada")
    base["cidade"] = base["cidade"].apply(
        lambda nome: nome.strip() if isinstance(nome, str) and nome.strip() else "Cidade nao informada"
    )
    return (
        base.groupby("cidade")[["entregas", "peso", "valor"]]
        .sum()
        .sort_values(["valor", "entregas", "peso"], ascending=False)
        .reset_index()
        .rename(columns={"cidade": "colaborador"})
    )


def resumir_placas(df: pd.DataFrame) -> pd.DataFrame:
    base = df.copy()
    base["placa"] = base["placa"].fillna("Placa nao informada")
    base["placa"] = base["placa"].apply(
        lambda nome: nome.strip().upper() if isinstance(nome, str) and nome.strip() else "Placa nao informada"
    )
    return (
        base.groupby("placa")[["entregas", "peso", "valor"]]
        .sum()
        .sort_values(["peso", "valor", "entregas"], ascending=False)
        .reset_index()
        .rename(columns={"placa": "colaborador"})
    )


def build_metrics(summary: pd.DataFrame, *, total_entregas: float | None = None, total_peso: float | None = None, total_valor: float | None = None) -> str:
    total_entregas = total_entregas if total_entregas is not None else summary["entregas"].sum()
    total_peso_kg = total_peso if total_peso is not None else summary["peso"].sum()
    total_valor = total_valor if total_valor is not None else summary["valor"].sum()
    entregas_text = responsive_text(
        format_number(total_entregas, 0),
        format_compact_number(total_entregas, 1),
    )
    peso_text = responsive_text(
        f"{format_number(total_peso_kg, 3)} kg",
        f"{format_compact_number(total_peso_kg, 1)} kg",
    )
    valor_text = responsive_text(
        f"R$ {format_number(total_valor, 2)}",
        f"R$ {format_compact_number(total_valor, 1)}",
    )
    resumo_text = (
        f"Peso: {responsive_text(format_number(total_peso_kg, 3), format_compact_number(total_peso_kg, 1))} kg "
        f"&bull; Valor: R$ {responsive_text(format_number(total_valor, 2), format_compact_number(total_valor, 1))}"
    )

    metric_cards = [
        (
            "metric-total",
            "Total de entregas",
            entregas_text,
            resumo_text,
        ),
        (
            "metric-primary",
            "Peso total (kg)",
            peso_text,
            "Somatorio do periodo",
        ),
        (
            "metric-success",
            "Valor faturado (R$)",
            valor_text,
            "Somatorio do periodo",
        ),
    ]

    cards_html = "\n".join(
        f"""      <article class="metric-card {css_class}">
        <span class="metric-title">{title}</span>
        <strong class="metric-value">{value}</strong>
        <span class="metric-sub">{subtitle}</span>
      </article>"""
        for css_class, title, value, subtitle in metric_cards
    )

    return f"""    <div class="metrics-grid">
{cards_html}
    </div>"""


def build_podium(summary: pd.DataFrame, photo_map: dict[str, str], *, sort_columns: list[str] | None = None) -> str:
    sort_columns = sort_columns or DEFAULT_SORT_COLUMNS
    top3 = summary.sort_values(sort_columns, ascending=False).head(3)
    if top3.empty:
        return ""

    podium_map = {rank: row for rank, row in enumerate(top3.itertuples(index=False), start=1)}
    slots = [
        (2, "podium-card podium-second"),
        (1, "podium-card podium-first"),
        (3, "podium-card podium-third"),
    ]

    cards: list[str] = []
    for position, css_class in slots:
        row = podium_map.get(position)
        if row is None:
            cards.append(
                f"""      <article class="{css_class} podium-empty">
        <div class="podium-placeholder">Disponivel</div>
      </article>"""
            )
            continue

        entregas = responsive_text(format_quantity(row.entregas), format_compact_number(row.entregas, 1))
        peso = responsive_text(format_number(row.peso, 2), format_compact_number(row.peso, 1))
        valor = responsive_text(format_number(row.valor, 2), format_compact_number(row.valor, 1))
        nome_original = str(row.colaborador).strip()
        nome = nome_original.title()
        photo_src = lookup_photo(nome_original, photo_map)
        if photo_src:
            avatar = f'<div class="podium-avatar has-photo"><img src="{photo_src}" alt="{nome}"></div>'
        else:
            avatar = f'<div class="podium-avatar">{get_initials(nome)}</div>'
        cards.append(
            f"""      <article class="{css_class}">
        <div class="podium-medal">#{position}</div>
        {avatar}
        <h3>{nome}</h3>
        <p class="podium-value">Entregas: <strong>{entregas}</strong></p>
        <p class="podium-value">Peso total: <strong>{peso} kg</strong></p>
        <p class="podium-value">Valor total: <strong>R$ {valor}</strong></p>
      </article>"""
        )

    return f"""    <div class="podium">
{chr(10).join(cards)}
    </div>"""


def build_ranking_table(summary: pd.DataFrame, *, name_label: str = "Colaborador") -> str:
    linhas: list[str] = []
    for rank, row in enumerate(summary.itertuples(index=False), start=1):
        classe = " class=\"is-top\"" if rank <= 3 else ""
        nome = row.colaborador.title()
        linhas.append(
            f"""        <tr{classe}>
          <td data-label="Rank">{rank:02d}</td>
          <td data-label="{name_label}">{nome}</td>
          <td data-label="Entregas">{responsive_text(format_quantity(row.entregas), format_compact_number(row.entregas, 1))}</td>
          <td data-label="Peso (kg)">{responsive_text(format_number(row.peso, 2), format_compact_number(row.peso, 1))}</td>
          <td data-label="Valor (R$)">{responsive_text(f"R$ {format_number(row.valor, 2)}", format_compact_number(row.valor, 1))}</td>
        </tr>"""
        )

    corpo = "\n".join(linhas)
    return f"""    <div class="ranking-table">
      <table>
        <thead>
          <tr>
            <th data-short="RK">Rank</th>
            <th data-short="COLAB">{name_label}</th>
            <th data-short="ENT">Entregas</th>
            <th data-short="PESO">Peso (kg)</th>
            <th data-short="VALOR">Valor (R$)</th>
          </tr>
        </thead>
        <tbody>
{corpo}
        </tbody>
      </table>
    </div>"""


def build_section(
    title: str,
    summary: pd.DataFrame,
    *,
    show_metrics: bool = True,
    photo_map: dict[str, str],
    name_label: str = "Colaborador",
    sort_columns: list[str] | None = None,
    ranking_text: str | None = None,
) -> str:
    if summary.empty:
        return f"""  <section class="panel">
    <div class="section-heading">
      <h2>{title}</h2>
      <p>Nenhum registro encontrado.</p>
    </div>
  </section>"""

    sort_columns = sort_columns or DEFAULT_SORT_COLUMNS
    ranking_text = ranking_text or DEFAULT_RANKING_TEXT
    ordenado = summary.sort_values(sort_columns, ascending=False)
    metrics_block = build_metrics(ordenado) if show_metrics else ""

    return f"""  <section class="panel">
    <div class="section-heading">
      <h2>{title}</h2>
      <p>{ranking_text}</p>
    </div>
{metrics_block}
{build_podium(ordenado, photo_map, sort_columns=sort_columns)}
{build_ranking_table(ordenado, name_label=name_label)}
  </section>"""


def build_overall_summary(motoristas: pd.DataFrame, ajudantes: pd.DataFrame, photo_map: dict[str, str]) -> str:
    total_motoristas = motoristas["colaborador"].nunique() if not motoristas.empty else 0
    total_ajudantes = ajudantes["colaborador"].nunique() if not ajudantes.empty else 0

    motoristas_ord = motoristas.sort_values(COLAB_SORT_COLUMNS, ascending=False)
    ajudantes_ord = ajudantes.sort_values(COLAB_SORT_COLUMNS, ascending=False)

    top_motorista = motoristas_ord.iloc[0] if not motoristas_ord.empty else None
    top_ajudante = ajudantes_ord.iloc[0] if not ajudantes_ord.empty else None

    motorista_nome = top_motorista.colaborador.title() if top_motorista is not None else "Sem registros"
    motorista_entregas = (
        responsive_text(format_quantity(top_motorista.entregas), format_compact_number(top_motorista.entregas, 1))
        if top_motorista is not None
        else "-"
    )
    motorista_peso = (
        responsive_text(format_number(top_motorista.peso, 2), format_compact_number(top_motorista.peso, 1))
        if top_motorista is not None
        else "-"
    )
    motorista_valor = (
        responsive_text(format_number(top_motorista.valor, 2), format_compact_number(top_motorista.valor, 1))
        if top_motorista is not None
        else "-"
    )
    motorista_foto = lookup_photo(top_motorista.colaborador if top_motorista is not None else None, photo_map)
    motorista_avatar_html = f'<img class="summary-avatar" src="{motorista_foto}" alt="{motorista_nome}">' if motorista_foto else ""

    ajudante_nome = top_ajudante.colaborador.title() if top_ajudante is not None else "Sem registros"
    ajudante_entregas = (
        responsive_text(format_quantity(top_ajudante.entregas), format_compact_number(top_ajudante.entregas, 1))
        if top_ajudante is not None
        else "-"
    )
    ajudante_peso = (
        responsive_text(format_number(top_ajudante.peso, 2), format_compact_number(top_ajudante.peso, 1))
        if top_ajudante is not None
        else "-"
    )
    ajudante_valor = (
        responsive_text(format_number(top_ajudante.valor, 2), format_compact_number(top_ajudante.valor, 1))
        if top_ajudante is not None
        else "-"
    )
    ajudante_foto = lookup_photo(top_ajudante.colaborador if top_ajudante is not None else None, photo_map)
    ajudante_avatar_html = f'<img class="summary-avatar" src="{ajudante_foto}" alt="{ajudante_nome}">' if ajudante_foto else ""

    return f"""  <section class="panel summary-panel">
    <div class="section-heading">
      <h2>Resumo geral</h2>
      <p>Quantidade total de colaboradores e lideres em entregas.</p>
    </div>
    <div class="summary-grid">
      <article class="summary-card motoristas">
        <span class="summary-tag">Motoristas</span>
        <strong class="summary-count">{total_motoristas}</strong>
        <div class="summary-highlight">
          {motorista_avatar_html}
          <span>Top peso: <strong>{motorista_nome}</strong></span>
        </div>
        <p class="summary-detail">Peso: {motorista_peso} kg &bull; Entregas: {motorista_entregas} &bull; Valor: R$ {motorista_valor}</p>
      </article>
      <article class="summary-card ajudantes">
        <span class="summary-tag">Ajudantes</span>
        <strong class="summary-count">{total_ajudantes}</strong>
        <div class="summary-highlight">
          {ajudante_avatar_html}
          <span>Top peso: <strong>{ajudante_nome}</strong></span>
        </div>
        <p class="summary-detail">Peso: {ajudante_peso} kg &bull; Entregas: {ajudante_entregas} &bull; Valor: R$ {ajudante_valor}</p>
      </article>
    </div>
  </section>"""


def build_city_section(cidades: pd.DataFrame) -> str:
    if cidades.empty:
        return """  <section class="panel city-panel">
    <div class="section-heading">
      <h2>Mapa de dominancia por cidade</h2>
      <p>Nenhum registro encontrado.</p>
    </div>
  </section>"""

    ordenado = cidades.sort_values(["valor", "entregas", "peso"], ascending=False).reset_index(drop=True)

    return f"""  <section class="panel city-panel">
    <div class="section-heading">
      <h2>Mapa de dominancia por cidade</h2>
      <p>Bolhas proporcionais ao valor faturado, mostrando as cidades com maior participacao.</p>
    </div>
{build_ranking_table(ordenado, name_label="Cidade")}
  </section>"""


def build_dupla_section(
    titulo: str,
    ranking: pd.DataFrame,
    name_label: str,
    *,
    sort_columns: list[str] | None = None,
    ranking_text: str | None = None,
) -> str:
    if ranking.empty:
        return f"""  <section class="panel">
    <div class="section-heading">
      <h2>{titulo}</h2>
      <p>Nenhum registro encontrado.</p>
    </div>
  </section>"""
    sort_columns = sort_columns or COLAB_SORT_COLUMNS
    ranking_text = ranking_text or COLAB_RANKING_TEXT
    ordenado = ranking.sort_values(sort_columns, ascending=False)
    return f"""  <section class="panel">
    <div class="section-heading">
      <h2>{titulo}</h2>
      <p>{ranking_text}</p>
    </div>
{build_ranking_table(ordenado, name_label=name_label)}
  </section>"""


def render_dashboard(
    motoristas: pd.DataFrame,
    ajudantes: pd.DataFrame,
    clientes: pd.DataFrame,
    cidades: pd.DataFrame,
    month_options: list[tuple[str, str]],
    monthly_blocks: dict[str, dict[str, str]],
    total_entregas: float | None = None,
    total_peso: float | None = None,
    total_valor: float | None = None,
    periodo: Iterable[datetime] | None = None,
    photo_map: dict[str, str] | None = None,
) -> str:
    photo_map = photo_map or {}
    periodo_texto = ""
    if periodo is not None:
        valores = [p for p in periodo if pd.notna(p)]
        if valores:
            inicio = min(valores).strftime("%d/%m/%Y")
            fim = max(valores).strftime("%d/%m/%Y")
            periodo_texto = f"Periodo analisado: {inicio} - {fim}"

    gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
    resumo_geral = build_overall_summary(motoristas, ajudantes, photo_map)

    clientes_section = build_section("Clientes", clientes, show_metrics=False, photo_map=photo_map, name_label="Cliente")
    cidades_section = build_city_section(cidades)
    metric_overall = build_metrics(motoristas, total_entregas=total_entregas, total_peso=total_peso, total_valor=total_valor)
    default_month_key = month_options[0][0] if month_options else "all"
    if default_month_key not in monthly_blocks:
        default_month_key = "all"
    periodo_inicial = monthly_blocks.get(default_month_key, {}).get("periodo") or periodo_texto
    options_html = "\n".join(
        f'        <option value="{key}"{" selected" if key == default_month_key else ""}>{label}</option>'
        for key, label in month_options
    )

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8">
  <title>Ranking Colaboradores &#8211; Maiores Entregas JR Ferragens &amp; Madeiras</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    :root {{
      color-scheme: light;
      --bg-1: #f5f3ff;
      --bg-2: #eef2ff;
      --panel: rgba(255, 255, 255, 0.92);
      --text-main: #1f2937;
      --text-muted: #6b7280;
      --primary: #6c5ce7;
      --primary-dark: #4f46e5;
      --success: #28c997;
      --pink: #ec4899;
      --shadow: 0 25px 50px -20px rgba(76, 81, 191, 0.45);
      --radius-xl: 26px;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Poppins", "Segoe UI", Arial, sans-serif;
      background: #ffffff;
      color: var(--text-main);
      padding: 34px 18px 56px;
      width: 100%;
      max-width: 100vw;
      overflow-x: hidden;
    }}
    main {{
      width: 100%;
      max-width: 1180px;
      margin: 0 auto;
      display: grid;
      gap: 28px;
    }}
    header.page-header {{
      background: var(--panel);
      border-radius: var(--radius-xl);
      padding: 28px 32px;
      box-shadow: var(--shadow);
      display: flex;
      flex-direction: column;
      gap: 12px;
    }}
    .page-header-top {{
      display: flex;
      align-items: center;
      gap: 18px;
      flex-wrap: wrap;
    }}
    .brand-logo {{
      width: auto;
      height: 58px;
      max-width: min(260px, 56vw);
      object-fit: contain;
    }}
    header.page-header h1 {{
      margin: 0;
      flex: 1 1 320px;
      min-width: 0;
      font-size: clamp(1.55rem, 3.6vw, 2.35rem);
      line-height: 1.18;
      color: var(--primary-dark);
      letter-spacing: -0.5px;
      overflow-wrap: anywhere;
    }}
    header.page-header p {{
      margin: 0;
      color: var(--text-muted);
    }}
    .page-meta {{
      display: flex;
      gap: 18px;
      flex-wrap: wrap;
      color: var(--text-muted);
      font-size: 0.95rem;
    }}
    .page-meta span {{
      min-width: 0;
      overflow-wrap: anywhere;
    }}
    .panel {{
      background: var(--panel);
      border-radius: var(--radius-xl);
      padding: clamp(18px, 2.8vw, 32px) clamp(16px, 2.4vw, 28px);
      backdrop-filter: blur(20px);
      box-shadow: var(--shadow);
      display: flex;
      flex-direction: column;
      gap: 24px;
      min-width: 0;
    }}
    .section-heading {{
      display: flex;
      flex-direction: column;
      gap: 6px;
    }}
    .section-heading h2 {{
      margin: 0;
      font-size: 1.65rem;
      letter-spacing: -0.3px;
    }}
    .section-heading p {{
      margin: 0;
      color: var(--text-muted);
      font-size: 0.95rem;
    }}
    .value-compact {{
      display: none;
    }}
    .value-full,
    .value-compact {{
      white-space: nowrap;
      font-variant-numeric: tabular-nums;
    }}
    .metrics-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: 18px;
    }}
    .metric-card {{
      border-radius: 22px;
      padding: 22px 24px;
      color: #ffffff;
      display: flex;
      flex-direction: column;
      gap: 10px;
      min-height: 140px;
      min-width: 0;
      position: relative;
      overflow: hidden;
      box-shadow: var(--shadow);
    }}
    .metric-title {{
      font-size: 0.78rem;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      opacity: 0.85;
    }}
    .metric-value {{
      font-size: 1.9rem;
      line-height: 1.2;
    }}
    .metric-sub {{
      font-size: 0.9rem;
      opacity: 0.9;
    }}
    .metric-card.metric-total {{
      background: linear-gradient(135deg, #818cf8, #4338ca);
    }}
    .metric-card.metric-primary {{
      background: linear-gradient(135deg, #6c5ce7, #4f46e5);
    }}
    .metric-card.metric-success {{
      background: linear-gradient(135deg, #22d3ee, #0ea5e9);
    }}
    .metric-card.metric-value {{
      background: linear-gradient(135deg, #f472b6, #ec4899);
    }}
    .summary-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
      gap: 20px;
      margin-top: 8px;
    }}
    .summary-card {{
      border-radius: 24px;
      padding: 24px;
      color: #ffffff;
      display: grid;
      gap: 10px;
      box-shadow: var(--shadow);
      position: relative;
      overflow: hidden;
      min-height: 160px;
    }}
    .summary-card.motoristas {{
      background: linear-gradient(135deg, #6366f1, #4338ca);
    }}
    .summary-card.ajudantes {{
      background: linear-gradient(135deg, #f97316, #ea580c);
    }}
    .summary-tag {{
      font-size: 0.78rem;
      text-transform: uppercase;
      letter-spacing: 0.1em;
      opacity: 0.8;
    }}
    .summary-count {{
      font-size: 2.5rem;
      font-weight: 700;
      line-height: 1.1;
    }}
    .summary-highlight {{
      font-size: 0.95rem;
      display: flex;
      align-items: center;
      gap: 12px;
    }}
    .summary-highlight strong {{
      font-weight: 700;
    }}
    .summary-avatar {{
      width: 48px;
      height: 48px;
      border-radius: 50%;
      object-fit: cover;
      border: 2px solid rgba(255, 255, 255, 0.45);
      box-shadow: 0 6px 14px -8px rgba(0, 0, 0, 0.45);
    }}
    .summary-detail {{
      margin: 0;
      font-size: 0.9rem;
      opacity: 0.9;
    }}
    .podium {{
      margin: 30px 0 16px;
      display: flex;
      justify-content: center;
      align-items: flex-end;
      gap: 20px;
      flex-wrap: wrap;
    }}
    .podium-card {{
      width: clamp(180px, 26vw, 240px);
      border-radius: 28px;
      padding: 26px 22px;
      text-align: center;
      position: relative;
      color: #ffffff;
      box-shadow: var(--shadow);
    }}
    .podium-first {{
      background: linear-gradient(135deg, #fde68a, #f59e0b);
      color: #3b2f0b;
    }}
    .podium-second {{
      background: linear-gradient(135deg, #cad5ff, #94a3b8);
      color: #1f2937;
    }}
    .podium-third {{
      background: linear-gradient(135deg, #fcd34d, #b45309);
      color: #3b2907;
    }}
    .podium-empty {{
      background: rgba(255, 255, 255, 0.6);
      color: var(--text-muted);
      border: 2px dashed rgba(148, 163, 184, 0.4);
      box-shadow: none;
    }}
    .podium-placeholder {{
      font-weight: 600;
      margin: 36px 0;
    }}
    .podium-medal {{
      position: absolute;
      top: -16px;
      right: 18px;
      padding: 8px 12px;
      border-radius: 999px;
      font-weight: 700;
      letter-spacing: 0.05em;
      background: rgba(255, 255, 255, 0.28);
    }}
    .podium-avatar {{
      width: 94px;
      height: 94px;
      margin: 10px auto 18px;
      border-radius: 50%;
      display: grid;
      place-items: center;
      font-size: 1.7rem;
      font-weight: 700;
      background: rgba(255, 255, 255, 0.28);
      border: 4px solid rgba(255, 255, 255, 0.38);
    }}
    .podium-avatar.has-photo {{
      background: none;
      border: none;
      display: flex;
      justify-content: center;
      align-items: center;
      padding: 0;
    }}
    .podium-avatar.has-photo img {{
      width: 94px;
      height: 94px;
      border-radius: 50%;
      object-fit: cover;
      border: 4px solid rgba(255, 255, 255, 0.55);
      box-shadow: 0 8px 18px -10px rgba(0, 0, 0, 0.4);
    }}
    .podium h3 {{
      margin: 0 0 10px;
      font-size: 1.15rem;
    }}
    .podium-value {{
      margin: 0;
      font-size: 0.95rem;
    }}
    .ranking-table {{
      background: rgba(255, 255, 255, 0.95);
      border-radius: 22px;
      box-shadow: var(--shadow);
      padding: 18px 22px;
      overflow-x: auto;
      max-height: 440px;
      overflow-y: auto;
      width: 100%;
      max-width: 100%;
      min-width: 0;
    }}
    .ranking-table table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 520px;
    }}
    .ranking-table thead th {{
      text-align: left;
      font-size: 0.75rem;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      color: var(--text-muted);
      padding: 12px 16px;
      border-bottom: 1px solid rgba(99, 102, 241, 0.12);
      white-space: nowrap;
    }}
    .ranking-table tbody td {{
      padding: 14px 16px;
      font-size: 0.95rem;
      color: var(--text-main);
      border-bottom: 1px solid rgba(99, 102, 241, 0.08);
    }}
    .ranking-table tbody tr:last-child td {{
      border-bottom: none;
    }}
    .ranking-table tbody td:nth-child(1) {{
      font-weight: 700;
      color: var(--primary);
    }}
    .ranking-table tbody td:nth-child(2) {{
      white-space: normal;
      word-break: break-word;
    }}
    .ranking-table tbody td:nth-child(3),
    .ranking-table tbody td:nth-child(4),
    .ranking-table tbody td:nth-child(5) {{
      text-align: right;
      font-variant-numeric: tabular-nums;
    }}
    .ranking-table tbody tr.is-top {{
      font-weight: 600;
      background: linear-gradient(90deg, rgba(108, 92, 231, 0.16), rgba(79, 70, 229, 0.05));
    }}
    .ranking-table tbody tr:hover {{
      background: rgba(243, 244, 255, 0.65);
    }}
    .city-grid-cards {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
      gap: 14px;
      margin-top: 12px;
    }}
    .city-card {{
      background: rgba(255,255,255,0.92);
      border-radius: 16px;
      padding: 14px 16px;
      box-shadow: 0 25px 40px -28px rgba(76,81,191,0.55);
      display: grid;
      grid-template-columns: auto 1fr;
      gap: 10px;
      align-items: center;
    }}
    .city-badge {{
      min-width: 46px;
      height: 32px;
      border-radius: 999px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      background: linear-gradient(135deg, #818cf8, #4f46e5);
      color: #ffffff;
      font-weight: 700;
      font-variant-numeric: tabular-nums;
      box-shadow: 0 10px 20px -14px rgba(79, 70, 229, 0.6);
    }}
    .city-name {{
      font-weight: 700;
      margin: 0;
      color: #111827;
    }}
    .city-meta {{
      margin: 2px 0 0;
      color: var(--text-muted);
      font-size: 0.92rem;
    }}
    .city-legend {{
      margin-top: 12px;
    }}
    .city-legend ul {{
      list-style: none;
      padding: 0;
      margin: 0;
    }}
    .filter-label {{
      font-weight: 600;
      color: var(--text-muted);
    }}
    .filter-select {{
      padding: 10px 12px;
      border-radius: 10px;
      border: 1px solid rgba(99, 102, 241, 0.2);
      box-shadow: 0 10px 24px -18px rgba(76,81,191,0.55);
      margin-left: 8px;
      min-width: 160px;
      background: #ffffff;
      font-size: 16px;
    }}
    .filter-group {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      flex-wrap: wrap;
      margin-bottom: 10px;
    }}
    footer {{
      text-align: center;
      color: var(--text-muted);
      font-size: 0.85rem;
    }}
    @media (max-width: 980px) {{
      main {{
        gap: 22px;
      }}
      .page-header-top {{
        align-items: flex-start;
      }}
      .podium {{
        justify-content: flex-start;
      }}
      .podium-card {{
        width: min(280px, 100%);
      }}
    }}
    @media (max-width: 720px) {{
      body {{
        padding: 18px 10px 34px;
      }}
      main {{
        gap: 16px;
      }}
      header.page-header {{
        padding: 14px 12px;
      }}
      .page-header-top {{
        width: 100%;
        display: grid;
        grid-template-columns: auto minmax(0, 1fr);
        align-items: center;
        gap: 10px;
      }}
      header.page-header h1 {{
        flex: 1 1 auto;
        min-width: 0;
      }}
      .brand-logo {{
        height: 34px;
      }}
      .page-meta {{
        display: grid;
        grid-template-columns: 1fr;
        gap: 6px;
        font-size: 0.8rem;
      }}
      .section-heading h2 {{
        font-size: 1.2rem;
      }}
      .metrics-grid {{
        grid-template-columns: 1fr;
        gap: 10px;
      }}
      .metric-card {{
        min-height: 0;
        padding: 14px 12px;
      }}
      .metric-title {{
        font-size: 0.7rem;
      }}
      .metric-value {{
        font-size: 1.32rem;
      }}
      .metric-sub {{
        font-size: 0.82rem;
      }}
      .summary-grid {{
        grid-template-columns: 1fr;
        gap: 10px;
      }}
      .summary-card {{
        min-height: 0;
        padding: 14px 12px;
      }}
      .summary-count {{
        font-size: 2rem;
      }}
      .podium {{
        margin: 8px 0;
        gap: 8px;
      }}
      .podium-card {{
        width: 100%;
        max-width: none;
        padding: 14px 10px;
      }}
      .podium-avatar,
      .podium-avatar.has-photo img {{
        width: 64px;
        height: 64px;
      }}
      .podium h3 {{
        font-size: 1rem;
      }}
      .podium-value {{
        font-size: 0.85rem;
      }}
      .filter-group {{
        display: grid;
        gap: 6px;
        width: 100%;
      }}
      .filter-select {{
        width: 100%;
        margin-left: 0;
      }}
      .ranking-table {{
        padding: 6px 6px;
        max-height: 210px;
        overflow-y: auto;
        overflow-x: hidden;
        -webkit-overflow-scrolling: touch;
      }}
      .ranking-table table {{
        min-width: 100%;
        width: 100%;
        table-layout: fixed;
      }}
      .ranking-table thead th,
      .ranking-table tbody td {{
        padding: 8px 10px;
        font-size: 0.82rem;
        white-space: nowrap;
      }}
      .ranking-table thead th {{
        font-size: 0;
        letter-spacing: 0;
        text-transform: none;
        padding-top: 6px;
        padding-bottom: 6px;
      }}
      .ranking-table thead th::after {{
        content: attr(data-short);
        font-size: 0.7rem;
        font-weight: 700;
        letter-spacing: 0.02em;
        color: var(--text-muted);
      }}
      .ranking-table thead th:nth-child(1),
      .ranking-table tbody td:nth-child(1) {{
        width: 9%;
      }}
      .ranking-table thead th:nth-child(2),
      .ranking-table tbody td:nth-child(2) {{
        width: 34%;
      }}
      .ranking-table thead th:nth-child(3),
      .ranking-table tbody td:nth-child(3) {{
        width: 13%;
      }}
      .ranking-table thead th:nth-child(4),
      .ranking-table tbody td:nth-child(4) {{
        width: 20%;
      }}
      .ranking-table thead th:nth-child(5),
      .ranking-table tbody td:nth-child(5) {{
        width: 24%;
      }}
      .ranking-table tbody td:nth-child(2) {{
        overflow: hidden;
        text-overflow: ellipsis;
      }}
      .ranking-table tbody td:nth-child(5) {{
        overflow: hidden;
        text-overflow: ellipsis;
      }}
    }}
    @media (max-width: 420px) {{
      header.page-header h1 {{
        font-size: 1.3rem;
      }}
      .section-heading p {{
        font-size: 0.82rem;
      }}
      .ranking-table {{
        max-height: 190px;
      }}
      .ranking-table table {{
        min-width: 100%;
      }}
      .ranking-table thead th {{
        font-size: 0;
        padding: 6px 6px;
      }}
      .ranking-table tbody td {{
        font-size: 0.76rem;
        padding: 8px 6px;
      }}
      .ranking-table thead th::after {{
        font-size: 0.62rem;
      }}
    }}
    @media (max-width: 520px) {{
      .value-full {{
        display: none;
      }}
      .value-compact {{
        display: inline;
      }}
    }}
  </style>
</head>
<body>
  <main>
    <header class="page-header">
      <div class="page-header-top">
        <img class="brand-logo" src="logo-jr.png" alt="Logo JR Ferragens e Madeiras">
        <h1>Ranking Colaboradores &#8211; Maiores Entregas JR Ferragens &amp; Madeiras</h1>
      </div>
     <div class="page-meta">
        <span>Atualizado em: {gerado_em}</span>
        <span id="periodo-texto">{periodo_inicial}</span>
      </div>
    </header>
    <div class="filter-group">
      <label class="filter-label" for="global-month-filter">Filtrar por mes:</label>
      <select id="global-month-filter" class="filter-select">
        <option value="all"{" selected" if default_month_key == "all" else ""}>Todos</option>
{options_html}
      </select>
    </div>
  <section class="panel panel-metrics" id="panel-metrics">
{monthly_blocks[default_month_key]["metrics"]}
  </section>
  <div id="section-motoristas">
{monthly_blocks[default_month_key]["motoristas"]}
  </div>
  <div id="section-ajudantes">
{monthly_blocks[default_month_key]["ajudantes"]}
  </div>
  <div id="section-clientes">
{monthly_blocks[default_month_key]["clientes"]}
  </div>
  <div id="section-placas">
{monthly_blocks[default_month_key]["placas"]}
  </div>
  <div id="section-cidades">
{monthly_blocks[default_month_key]["cidades"]}
  </div>
  <div id="section-mot-cidade">
{monthly_blocks[default_month_key]["mot_cidade"]}
  </div>
  <div id="section-aj-cidade">
{monthly_blocks[default_month_key]["aj_cidade"]}
  </div>
  <div id="section-mot-cliente">
{monthly_blocks[default_month_key]["mot_cliente"]}
  </div>
  <div id="section-aj-cliente">
{monthly_blocks[default_month_key]["aj_cliente"]}
  </div>
  <div id="section-resumo">
{monthly_blocks[default_month_key]["resumo"]}
  </div>
    <footer>
      Dashboard gerado automaticamente a partir da planilha "{EXCEL_NAME}".
    </footer>
  </main>
</body>
<script>
  (function() {{
    const data = {json.dumps(monthly_blocks, ensure_ascii=False)};
    const select = document.getElementById("global-month-filter");
    const periodSpan = document.getElementById("periodo-texto");
    const targets = {{
      metrics: document.getElementById("panel-metrics"),
      motoristas: document.getElementById("section-motoristas"),
      ajudantes: document.getElementById("section-ajudantes"),
      clientes: document.getElementById("section-clientes"),
      placas: document.getElementById("section-placas"),
      cidades: document.getElementById("section-cidades"),
      mot_cidade: document.getElementById("section-mot-cidade"),
      aj_cidade: document.getElementById("section-aj-cidade"),
      mot_cliente: document.getElementById("section-mot-cliente"),
      aj_cliente: document.getElementById("section-aj-cliente"),
      resumo: document.getElementById("section-resumo"),
    }};
    function render(key) {{
      const block = data[key] || data["all"];
      if (!block) return;
      targets.metrics.innerHTML = block.metrics;
      targets.motoristas.innerHTML = block.motoristas;
      targets.ajudantes.innerHTML = block.ajudantes;
      targets.clientes.innerHTML = block.clientes;
      targets.placas.innerHTML = block.placas;
      targets.cidades.innerHTML = block.cidades;
      targets.mot_cidade.innerHTML = block.mot_cidade;
      targets.aj_cidade.innerHTML = block.aj_cidade;
      targets.mot_cliente.innerHTML = block.mot_cliente;
      targets.aj_cliente.innerHTML = block.aj_cliente;
      targets.resumo.innerHTML = block.resumo;
      if (periodSpan) periodSpan.textContent = block.periodo || "";
    }}
    if (select) {{
      select.addEventListener("change", () => render(select.value));
    }}
    render(select ? select.value : "all");
  }})();
</script>
</body>
</html>
"""


def main() -> None:
    base_path = Path(__file__).resolve().parent
    excel_path = base_path / EXCEL_NAME
    if not excel_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {excel_path}")

    df = load_planilha(excel_path)
    motoristas, ajudantes = resumir_colaboradores(df)
    clientes = resumir_clientes(df)
    cidades = resumir_cidades(df)
    # Dados por cidade para filtro mensal
    city_rows: list[dict[str, object]] = []
    month_keys: dict[str, str] = {}
    for _, row in df.iterrows():
        data = row.get("data")
        cidade = row.get("cidade")
        if pd.isna(data) or not isinstance(cidade, str):
            continue
        mes_key = data.strftime("%Y-%m")
        mes_label = data.strftime("%m/%Y")
        month_keys[mes_key] = mes_label
        city_rows.append(
            {
                "mes": mes_key,
                "cidade": cidade.strip().title(),
                "entregas": float(row.get("entregas", 0) or 0),
                "peso": float(row.get("peso", 0) or 0),
                "valor": float(row.get("valor", 0) or 0),
            }
        )
    month_options = sorted(month_keys.items(), reverse=True)
    # Fotos: prioridade para pasta local (PHOTO_INPUT_DIR) e, se existir, imagens embutidas na planilha
    local_photos = load_local_photos(base_path / PHOTO_INPUT_DIR)
    excel_photos = collect_photos(excel_path)
    photo_map = {**excel_photos, **local_photos}
    total_entregas = df["entregas"].sum()
    total_peso = df["peso"].sum()
    total_valor = df["valor"].sum()

    def period_text(subset: pd.DataFrame) -> str:
        valores = [p for p in subset["data"] if pd.notna(p)]
        if not valores:
            return ""
        return f"Periodo analisado: {min(valores).strftime('%d/%m/%Y')} - {max(valores).strftime('%d/%m/%Y')}"

    monthly_blocks: dict[str, dict[str, str]] = {}

    def compute_blocks(df_subset: pd.DataFrame, label_periodo: str) -> dict[str, str]:
        mot, aj = resumir_colaboradores(df_subset)
        cli = resumir_clientes(df_subset)
        pla = resumir_placas(df_subset)
        cid = resumir_cidades(df_subset)
        mot_cidade = ranking_motorista_por(df_subset, "cidade")
        aj_cidade = ranking_ajudante_por(df_subset, "cidade")
        mot_cliente = ranking_motorista_por(df_subset, "cliente")
        aj_cliente = ranking_ajudante_por(df_subset, "cliente")
        totals_ent = df_subset["entregas"].sum()
        totals_peso = df_subset["peso"].sum()
        totals_valor = df_subset["valor"].sum()
        return {
            "metrics": build_metrics(mot, total_entregas=totals_ent, total_peso=totals_peso, total_valor=totals_valor),
            "motoristas": build_section(
                "Motoristas",
                mot,
                show_metrics=False,
                photo_map=photo_map,
                sort_columns=COLAB_SORT_COLUMNS,
                ranking_text=COLAB_RANKING_TEXT,
            ),
            "ajudantes": build_section(
                "Ajudantes",
                aj,
                show_metrics=False,
                photo_map=photo_map,
                sort_columns=COLAB_SORT_COLUMNS,
                ranking_text=COLAB_RANKING_TEXT,
            ),
            "clientes": build_section("Clientes", cli, show_metrics=False, photo_map=photo_map, name_label="Cliente"),
            "placas": build_section("Placas", pla, show_metrics=False, photo_map=photo_map, name_label="Placa"),
            "cidades": build_city_section(cid),
            "mot_cidade": build_dupla_section("Motoristas por cidade", mot_cidade, "Motorista — Cidade"),
            "aj_cidade": build_dupla_section("Ajudantes por cidade", aj_cidade, "Ajudante — Cidade"),
            "mot_cliente": build_dupla_section("Motoristas por cliente", mot_cliente, "Motorista — Cliente"),
            "aj_cliente": build_dupla_section("Ajudantes por cliente", aj_cliente, "Ajudante — Cliente"),
            "resumo": build_overall_summary(mot, aj, photo_map),
            "periodo": label_periodo,
        }

    # All
    monthly_blocks["all"] = compute_blocks(df, period_text(df))

    # Per month
    for key, _label in month_options:
        subset = df[df["data"].dt.strftime("%Y-%m") == key]
        monthly_blocks[key] = compute_blocks(subset, period_text(subset))

    html = render_dashboard(
        motoristas,
        ajudantes,
        clientes,
        cidades,
        month_options,
        monthly_blocks,
        total_entregas,
        total_peso,
        total_valor,
        df["data"].dropna().tolist(),
        photo_map,
    )
    output_path = base_path / OUTPUT_NAME
    output_path.write_text(html, encoding="utf-8")
    print(f"Dashboard gerado em: {output_path}")


if __name__ == "__main__":
    main()
